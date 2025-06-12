import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle


def highlight_top_10(csv_file):
    """Create an Excel file with top 10% highlighted in green."""
    df = pd.read_csv(csv_file)
    wb = Workbook()
    ws = wb.active
    ws.title = "Range"

    # Write headers and rows
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))

    highlight_cols = ["PERIOD_M1", "PERIOD_M5", "PERIOD_M15", "PERIOD_M30"]
    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    last_row = len(df) + 1  # 1-based with header

    for col_name in highlight_cols:
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name) + 1
            col_letter = get_column_letter(col_idx)
            rule = Rule(type="top10", rank=10, percent=True)
            rule.dxf = DifferentialStyle(fill=fill)
            rule.stopIfTrue = False
            ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{last_row}", rule)

    output_file = Path(csv_file).with_name("Range_highlighted.xlsx")
    wb.save(output_file)
    print(f"Saved highlighted file to {output_file}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python highlight_range_csv.py path_to_Range.csv")
        sys.exit(1)
    highlight_top_10(sys.argv[1])
